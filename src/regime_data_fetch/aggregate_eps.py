from __future__ import annotations

import datetime as dt
import json
from dataclasses import asdict, dataclass
from pathlib import Path
import urllib.error
import urllib.request

import pandas as pd
from openpyxl import load_workbook

from regime_data_fetch.acquisition_store import AcquisitionStore


SOURCE_NAME = "S&P Global aggregate forward EPS workbook"
SOURCE_URL = "https://www.spglobal.com/spdji/en/documents/additional-material/sp-500-eps-est.xlsx"
SHEET_NAME = "ESTIMATES&PEs"
WAYBACK_CDX_URL = "https://web.archive.org/cdx/search/cdx"

# Weekly-snapshot accumulator (Log #48 closure path). Each weekly run of
# `run_aggregate_eps_fetch` appends the workbook's current snapshot to this
# parquet, deduped by observation_date. Once >= 4 distinct weekly rows have
# accumulated, `compute_eps_revision_direction_4w` produces a non-NaN
# revision series and the §2B `earnings_expansion` / `earnings_contraction`
# labels unlock. The single S&P workbook only exposes quarterly history +
# one current point, so weekly granularity can only be built by
# accumulating one current-snapshot row per weekly fetch.
WEEKLY_HISTORY_FILENAME = "sp500_eps_weekly_history.parquet"
# Spec §2B: revision direction over 4 weeks. 4 rows back in the weekly-sorted
# accumulator history (one row per weekly fetch).
EPS_REVISION_LOOKBACK_WEEKS = 4
# Output sub-directory + Wayback timeline filenames (shared by the live
# fetch, the Wayback backfill, and the accumulator-seeding bridge).
EPS_DIR_NAME = "aggregate_forward_eps"
WAYBACK_DIR_NAME = "aggregate_forward_eps_wayback"
WAYBACK_TIMELINE_FILENAME = "sp500_eps_wayback_timeline.parquet"


class AggregateEPSFetchError(RuntimeError):
    pass


@dataclass(frozen=True)
class AggregateEPSSnapshot:
    observation_date: dt.date
    observation_label: str
    forward_estimate_label: str | None
    forward_estimate_value: float | None
    estimate_2025e: float | None
    estimate_q4_2025e: float | None
    estimate_2026e: float | None
    price: float | None
    pe_2025e: float | None
    pe_2026e: float | None
    change_vs_prior_observation_2025e: float | None
    change_vs_prior_observation_q4_2025e: float | None
    change_vs_prior_observation_2026e: float | None
    change_vs_prior_observation_price: float | None
    change_vs_prior_observation_pe_2025e: float | None
    change_vs_prior_observation_pe_2026e: float | None


@dataclass(frozen=True)
class ParsedAggregateEPSWorkbook:
    workbook_as_of_date: dt.date
    public_files_discontinued: bool
    historical_snapshots: list[AggregateEPSSnapshot]
    current_snapshot: AggregateEPSSnapshot


def parse_sp500_eps_workbook(workbook_path: Path) -> ParsedAggregateEPSWorkbook:
    if workbook_path.suffix.lower() == ".xls":
        return _parse_legacy_sp500_eps_workbook(workbook_path)

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise AggregateEPSFetchError(f"Workbook missing expected sheet {SHEET_NAME!r}")

    ws = wb[SHEET_NAME]
    workbook_as_of_date = _extract_workbook_as_of_date(ws)
    public_files_discontinued = _extract_discontinued_flag(ws)
    table_start_row, header_labels = _find_observation_header_row(ws)
    current_changes = _find_current_change_row(ws, table_start_row)

    historical: list[AggregateEPSSnapshot] = []
    current_snapshot: AggregateEPSSnapshot | None = None
    for row_idx in range(table_start_row + 1, ws.max_row + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        first = row[0]
        if isinstance(first, dt.datetime):
            label_map = _build_observation_value_map(header_labels, row)
            historical.append(
                AggregateEPSSnapshot(
                    observation_date=first.date(),
                    observation_label="historical_quarter_end",
                    forward_estimate_label=_select_forward_estimate_label(header_labels),
                    forward_estimate_value=_select_forward_estimate_value(label_map, header_labels),
                    estimate_2025e=_value_for_exact_label(label_map, "2025E"),
                    estimate_q4_2025e=_value_for_exact_label(label_map, "Q4 2025E"),
                    estimate_2026e=_value_for_exact_label(label_map, "2026E"),
                    price=_value_for_price(label_map),
                    pe_2025e=_value_for_exact_label(label_map, "2025E P/E"),
                    pe_2026e=_value_for_pe(label_map, "2026"),
                    change_vs_prior_observation_2025e=None,
                    change_vs_prior_observation_q4_2025e=None,
                    change_vs_prior_observation_2026e=None,
                    change_vs_prior_observation_price=None,
                    change_vs_prior_observation_pe_2025e=None,
                    change_vs_prior_observation_pe_2026e=None,
                )
            )
            continue

        if isinstance(first, str) and first.strip().lower() == "current":
            label_map = _build_observation_value_map(header_labels, row)
            current_snapshot = AggregateEPSSnapshot(
                observation_date=workbook_as_of_date,
                observation_label="current",
                forward_estimate_label=_select_forward_estimate_label(header_labels),
                forward_estimate_value=_select_forward_estimate_value(label_map, header_labels),
                estimate_2025e=_value_for_exact_label(label_map, "2025E"),
                estimate_q4_2025e=_value_for_exact_label(label_map, "Q4 2025E"),
                estimate_2026e=_value_for_exact_label(label_map, "2026E"),
                price=_value_for_price(label_map),
                pe_2025e=_value_for_exact_label(label_map, "2025E P/E"),
                pe_2026e=_value_for_pe(label_map, "2026"),
                change_vs_prior_observation_2025e=current_changes[0],
                change_vs_prior_observation_q4_2025e=current_changes[1],
                change_vs_prior_observation_2026e=current_changes[2],
                change_vs_prior_observation_price=current_changes[3],
                change_vs_prior_observation_pe_2025e=current_changes[4],
                change_vs_prior_observation_pe_2026e=current_changes[5],
            )
            continue

        if current_snapshot is not None:
            break

    if not historical:
        raise AggregateEPSFetchError("Workbook contained no historical aggregate EPS snapshots")
    if current_snapshot is None:
        raise AggregateEPSFetchError("Workbook missing current aggregate EPS snapshot row")

    return ParsedAggregateEPSWorkbook(
        workbook_as_of_date=workbook_as_of_date,
        public_files_discontinued=public_files_discontinued,
        historical_snapshots=historical,
        current_snapshot=current_snapshot,
    )


_SPGLOBAL_EPS_MANUAL_REL_PATH = Path("spglobal_eps") / "sp-500-eps-est.xlsx"


def download_spglobal_eps_workbook(
    *,
    out_path: Path,
    source_url: str = SOURCE_URL,
    timeout_seconds: int = 60,
) -> Path:
    """Attempt to download the S&P Global aggregate forward-EPS workbook
    from the canonical public URL into ``out_path``.

    Cadence intent: the spdji workbook is published WEEKLY (typically Wed/Thu
    around the earnings revision cycle). Slice 5 §2B's deferred
    ``aggregate_forward_eps_revision_direction_4w`` predicate needs at least
    4 consecutive weekly observations to compute the rolling 4-week
    direction, so this fetcher is intended to run on a weekly schedule.

    Known issue: ``www.spglobal.com`` is served behind Akamai (AkamaiGHost)
    bot mitigation that returns HTTP 403 to direct HTTP requests including
    browser-User-Agent spoofs. ``run_aggregate_eps_auto_fetch`` handles that
    by trying a browser-backed download next, while still honoring an
    operator-staged workbook at ``data/raw/spglobal_eps/sp-500-eps-est.xlsx``.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(
        source_url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/126.0.0.0 Safari/537.36"
            ),
            "Accept": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                "application/octet-stream,*/*"
            ),
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            payload = response.read()
    except urllib.error.HTTPError as exc:
        if exc.code == 403:
            raise AggregateEPSFetchError(
                f"S&P spdji returned HTTP 403 (Akamai bot mitigation) for "
                f"{source_url}. The URL is browser-only; programmatic clients "
                f"are blocked. To complete the fetch: (1) open the URL in a "
                f"browser and download sp-500-eps-est.xlsx; (2) copy it to "
                f"data/raw/spglobal_eps/sp-500-eps-est.xlsx in this repo; "
                f"(3) re-run --fetch eps-spglobal-auto — the auto path "
                f"detects the manually-dropped file and parses it."
            ) from exc
        raise AggregateEPSFetchError(
            f"Failed to download S&P EPS workbook from {source_url}: {exc}"
        ) from exc
    except urllib.error.URLError as exc:
        raise AggregateEPSFetchError(
            f"Failed to download S&P EPS workbook from {source_url}: {exc}"
        ) from exc
    if not payload:
        raise AggregateEPSFetchError(
            f"S&P EPS workbook download from {source_url} returned empty payload"
        )
    out_path.write_bytes(payload)
    return out_path


def download_spglobal_eps_workbook_with_browser(
    *,
    out_path: Path,
    source_url: str = SOURCE_URL,
    timeout_ms: int = 120_000,
    user_data_dir: Path | None = None,
    executable_path: Path | None = None,
    headless: bool = True,
) -> Path:
    """Download the S&P workbook through a real browser session.

    This is the long-term fallback for Akamai-protected S&P downloads: direct
    HTTP remains first, then a scheduler can use a persistent browser profile
    that has already passed the provider's browser checks.
    """
    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise AggregateEPSFetchError(
            "Playwright is required for browser-backed EPS download. "
            "Install the browser extra and browser runtime, or pre-stage the workbook."
        ) from exc

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as playwright:
        launch_kwargs: dict[str, object] = {
            "headless": headless,
            "accept_downloads": True,
        }
        if executable_path is not None:
            launch_kwargs["executable_path"] = str(executable_path)

        browser = None
        context = None
        try:
            if user_data_dir is not None:
                context = playwright.chromium.launch_persistent_context(  # type: ignore[arg-type]
                    str(user_data_dir),
                    **launch_kwargs,  # type: ignore[arg-type]
                )
            else:
                browser = playwright.chromium.launch(
                    headless=headless,
                    executable_path=str(executable_path) if executable_path else None,
                )
                context = browser.new_context(accept_downloads=True)
            page = context.new_page()
            try:
                with page.expect_download(timeout=timeout_ms) as download_info:
                    page.goto(source_url, wait_until="domcontentloaded", timeout=timeout_ms)
                download = download_info.value
                download.save_as(out_path)
            except PlaywrightTimeoutError as exc:
                raise AggregateEPSFetchError(
                    "Browser-backed S&P EPS download did not produce a workbook download before timeout"
                ) from exc
        finally:
            if context is not None:
                context.close()
            if browser is not None:
                browser.close()
    return out_path


def run_aggregate_eps_auto_fetch(
    *,
    out_dir: Path,
    source_url: str = SOURCE_URL,
    acquisition_db_path: Path | None = None,
    artifact_store_root: str | Path | None = None,
    workbook_downloader=download_spglobal_eps_workbook,
    browser_downloader=download_spglobal_eps_workbook_with_browser,
    browser_user_data_dir: Path | None = None,
    browser_executable: Path | None = None,
    browser_headless: bool = True,
    browser_timeout_ms: int = 120_000,
) -> Path:
    """Fetch + parse the latest S&P aggregate-EPS workbook.

    Two-step resolution:
    1. If ``out_dir / spglobal_eps / sp-500-eps-est.xlsx`` already exists
       (operator manually downloaded — see
       ``download_spglobal_eps_workbook`` docstring for why), parse it
       directly. This is the canonical weekly cadence path:
         a. Each week, operator opens the spdji URL in a browser, downloads
            the .xlsx, copies it to data/raw/spglobal_eps/sp-500-eps-est.xlsx.
         b. Operator (or a scheduler) runs ``--fetch eps-spglobal-auto``,
            which detects the file and emits the same parquet + report
            artifacts as the manual ``--eps-workbook`` path.
    2. If the file is absent, try downloading from ``source_url`` directly.
       If direct download is blocked, try the browser-backed download path.

    Cadence: invoke weekly. Polling daily is wasteful — the workbook URL
    serves the same file between weekly publications.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = out_dir / _SPGLOBAL_EPS_MANUAL_REL_PATH
    if not workbook_path.exists():
        try:
            workbook_downloader(out_path=workbook_path, source_url=source_url)
        except AggregateEPSFetchError:
            browser_downloader(
                out_path=workbook_path,
                source_url=source_url,
                timeout_ms=browser_timeout_ms,
                user_data_dir=browser_user_data_dir,
                executable_path=browser_executable,
                headless=browser_headless,
            )
    return run_aggregate_eps_fetch(
        out_dir=out_dir,
        workbook_path=workbook_path,
        acquisition_db_path=acquisition_db_path,
        artifact_store_root=artifact_store_root,
    )


def append_weekly_eps_snapshot(
    *,
    eps_dir: Path,
    current_snapshot: AggregateEPSSnapshot,
) -> pd.DataFrame:
    """Append one weekly current-snapshot row to the accumulator parquet.

    Idempotent by ``observation_date``: re-running the same weekly workbook
    overwrites that date's row rather than double-counting it (the operator
    may re-run a fetch for the same week). Returns the full accumulated
    weekly-history DataFrame (sorted ascending by observation_date).

    Closes Log #48's "workbook snapshot path does not expose weekly time
    series" blocker — the weekly series is built by accumulating one row
    per weekly fetch rather than read from a single workbook.
    """
    history_path = eps_dir / WEEKLY_HISTORY_FILENAME
    new_row = pd.DataFrame(
        [
            {
                "observation_date": current_snapshot.observation_date,
                "observation_label": current_snapshot.observation_label,
                "forward_estimate_value": current_snapshot.forward_estimate_value,
                "source": SOURCE_NAME,
            }
        ]
    )
    if history_path.exists():
        existing = pd.read_parquet(history_path)
        # Drop any prior row for the same observation_date (idempotent
        # re-run), then append the fresh row.
        existing = existing[
            existing["observation_date"] != current_snapshot.observation_date
        ]
        combined = pd.concat([existing, new_row], ignore_index=True)
    else:
        combined = new_row
    combined = combined.sort_values("observation_date").reset_index(drop=True)
    combined.to_parquet(history_path, index=False)
    return combined


def seed_weekly_history_from_wayback_timeline(
    *,
    out_dir: Path,
    timeline_path: Path | None = None,
) -> pd.DataFrame:
    """Seed the weekly-history accumulator from a Wayback backfill timeline.

    ``run_wayback_aggregate_eps_fetch`` materialises historical workbook
    snapshots into ``sp500_eps_wayback_timeline.parquet`` but never feeds
    them into the weekly-history accumulator that
    ``compute_eps_revision_direction_4w`` reads. This bridges that gap: each
    timeline row becomes one accumulator row keyed by ``workbook_as_of_date``.

    Collapses the §2B `earnings_expansion` / `earnings_contraction`
    cold-start. Instead of waiting for more than ``EPS_REVISION_LOOKBACK_WEEKS``
    *live* weekly fetches to accumulate, a one-time Wayback backfill + seed
    pre-fills the accumulator and the 4-week revision series goes non-NaN
    immediately.

    Idempotent and live-safe: on an ``observation_date`` collision the
    EXISTING accumulator row wins — a live ``run_aggregate_eps_fetch`` row is
    authoritative over a Wayback-archived snapshot for the same date, and
    re-running the seed never clobbers live data. Returns the full merged
    weekly-history DataFrame, sorted ascending by ``observation_date``.
    """
    if timeline_path is None:
        timeline_path = out_dir / WAYBACK_DIR_NAME / WAYBACK_TIMELINE_FILENAME
    if not timeline_path.exists():
        raise AggregateEPSFetchError(
            f"No Wayback EPS timeline at {timeline_path}. Run "
            f"run_wayback_aggregate_eps_fetch first to materialise it."
        )
    timeline = pd.read_parquet(timeline_path)
    if timeline.empty:
        raise AggregateEPSFetchError(
            f"Wayback EPS timeline at {timeline_path} contained no rows"
        )

    seeded = pd.DataFrame(
        {
            "observation_date": timeline["workbook_as_of_date"],
            "observation_label": "wayback_backfill",
            "forward_estimate_value": timeline["forward_estimate_value"],
            "source": "wayback_machine",
        }
    )
    # A single workbook_as_of_date can appear under multiple Wayback
    # snapshots — keep the last (the timeline is sorted ascending by
    # snapshot_date / timestamp, so the last row is the freshest capture).
    seeded = seeded.drop_duplicates(subset=["observation_date"], keep="last")

    eps_dir = out_dir / EPS_DIR_NAME
    eps_dir.mkdir(parents=True, exist_ok=True)
    history_path = eps_dir / WEEKLY_HISTORY_FILENAME
    if history_path.exists():
        existing = pd.read_parquet(history_path)
        # Existing accumulator rows win on collision — a live fetch row is
        # authoritative over a Wayback snapshot for the same date.
        seeded = seeded[
            ~seeded["observation_date"].isin(existing["observation_date"])
        ]
        combined = pd.concat([existing, seeded], ignore_index=True)
    else:
        combined = seeded
    combined = combined.sort_values("observation_date").reset_index(drop=True)
    combined.to_parquet(history_path, index=False)
    return combined


def compute_eps_revision_direction_4w(weekly_history: pd.DataFrame) -> pd.Series:
    """v2 §2B `aggregate_forward_eps_revision_direction_4w` from the
    accumulated weekly history.

    ``revision_4w = (forward_eps[t] - forward_eps[t-4]) / forward_eps[t-4]``
    where ``t-4`` is 4 rows back in the weekly-sorted accumulator (one row
    per weekly fetch). The returned Series is indexed by
    ``observation_date``; values are NaN for the first
    ``EPS_REVISION_LOOKBACK_WEEKS`` rows (cold-start) and wherever the
    4-weeks-prior estimate is NaN or zero.

    Until the accumulator holds more than ``EPS_REVISION_LOOKBACK_WEEKS``
    weekly rows the entire series is NaN — the §2B earnings labels stay
    silent, which is the correct cold-start behaviour (V1 §2.7).
    """
    if weekly_history.empty:
        return pd.Series(
            dtype=float, name="aggregate_forward_eps_revision_direction_4w"
        )
    sorted_history = weekly_history.sort_values("observation_date").reset_index(
        drop=True
    )
    forward_eps = sorted_history["forward_estimate_value"].astype(float)
    prior = forward_eps.shift(EPS_REVISION_LOOKBACK_WEEKS)
    revision = (forward_eps - prior) / prior.where(prior != 0)
    revision.index = pd.DatetimeIndex(
        pd.to_datetime(sorted_history["observation_date"])
    )
    revision.name = "aggregate_forward_eps_revision_direction_4w"
    return revision


def run_aggregate_eps_fetch(
    *,
    out_dir: Path,
    workbook_path: Path,
    acquisition_db_path: Path | None = None,
    artifact_store_root: str | Path | None = None,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    store = AcquisitionStore(acquisition_db_path, artifact_store_root=artifact_store_root) if acquisition_db_path else None
    fetch_run = (
        store.start_fetch_run(
            fetch_type="aggregate_eps",
            params={
                "workbook_path": str(workbook_path),
            },
        )
        if store
        else None
    )

    try:
        if store and fetch_run:
            store.record_file_artifact(
                run_id=fetch_run.run_id,
                source_name=SOURCE_NAME,
                artifact_kind=f"{workbook_path.suffix.lower().lstrip('.')}_manual",
                source_identifier=str(workbook_path),
                file_path=workbook_path,
                timezone="America/New_York",
                license_note="Manually downloaded workbook; public files reported discontinued by source workbook",
                notes="Manual S&P aggregate EPS workbook snapshot",
            )

        parsed = parse_sp500_eps_workbook(workbook_path)

        rows = [*parsed.historical_snapshots, parsed.current_snapshot]
        df = pd.DataFrame(
            [
                {
                    "workbook_as_of_date": parsed.workbook_as_of_date,
                    "observation_date": row.observation_date,
                    "observation_label": row.observation_label,
                    "forward_estimate_label": row.forward_estimate_label,
                    "forward_estimate_value": row.forward_estimate_value,
                    "estimate_2025e": row.estimate_2025e,
                    "estimate_q4_2025e": row.estimate_q4_2025e,
                    "estimate_2026e": row.estimate_2026e,
                    "price": row.price,
                    "pe_2025e": row.pe_2025e,
                    "pe_2026e": row.pe_2026e,
                    "change_vs_prior_observation_2025e": row.change_vs_prior_observation_2025e,
                    "change_vs_prior_observation_q4_2025e": row.change_vs_prior_observation_q4_2025e,
                    "change_vs_prior_observation_2026e": row.change_vs_prior_observation_2026e,
                    "change_vs_prior_observation_price": row.change_vs_prior_observation_price,
                    "change_vs_prior_observation_pe_2025e": row.change_vs_prior_observation_pe_2025e,
                    "change_vs_prior_observation_pe_2026e": row.change_vs_prior_observation_pe_2026e,
                    "source": SOURCE_NAME,
                    "source_path": str(workbook_path),
                    "public_files_discontinued": parsed.public_files_discontinued,
                }
                for row in rows
            ]
        )

        eps_dir = out_dir / EPS_DIR_NAME
        eps_dir.mkdir(parents=True, exist_ok=True)
        parquet_path = eps_dir / "sp500_eps_snapshots.parquet"
        df.to_parquet(parquet_path, index=False)

        # Weekly-snapshot accumulator (Log #48 closure). Append this run's
        # current snapshot to the persistent weekly-history parquet, then
        # compute the 4-week revision direction. The revision series is
        # all-NaN until > EPS_REVISION_LOOKBACK_WEEKS weekly rows have
        # accumulated — the report's availability flag reflects that.
        weekly_history = append_weekly_eps_snapshot(
            eps_dir=eps_dir, current_snapshot=parsed.current_snapshot
        )
        weekly_history_path = eps_dir / WEEKLY_HISTORY_FILENAME
        revision_series = compute_eps_revision_direction_4w(weekly_history)
        revision_available = bool(revision_series.notna().any())

        current_dict = asdict(parsed.current_snapshot)
        current_dict["observation_date"] = parsed.current_snapshot.observation_date.isoformat()
        report = {
            "as_of_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
            "source": SOURCE_NAME,
            "source_url": SOURCE_URL,
            "source_path": str(workbook_path),
            "workbook_as_of_date": parsed.workbook_as_of_date.isoformat(),
            "public_files_discontinued": parsed.public_files_discontinued,
            "counts": {
                "historical_snapshots": len(parsed.historical_snapshots),
                "current_snapshots": 1,
                "weekly_history_rows": len(weekly_history),
            },
            "current_snapshot": current_dict,
            "limitations": {
                "aggregate_forward_eps_revision_direction_4w_available": revision_available,
                "reason": (
                    "Revision direction available — the weekly-snapshot accumulator "
                    f"holds {len(weekly_history)} rows (> {EPS_REVISION_LOOKBACK_WEEKS} "
                    "required for the 4-week lookback)."
                    if revision_available
                    else (
                        "The single S&P workbook exposes quarterly history plus one "
                        f"current snapshot. The weekly accumulator holds "
                        f"{len(weekly_history)} row(s); "
                        f"> {EPS_REVISION_LOOKBACK_WEEKS} weekly fetches are required "
                        "before the 4-week revision direction is non-NaN."
                    )
                ),
            },
            "paths": {
                "aggregate_eps_parquet": str(parquet_path),
                "aggregate_eps_weekly_history_parquet": str(weekly_history_path),
                "acquisition_db": str(acquisition_db_path) if acquisition_db_path else None,
            },
        }
        report_path = out_dir / "aggregate_eps_fetch_report.json"
        report_path.write_text(json.dumps(report, indent=2))

        if store and fetch_run:
            store.record_output(
                run_id=fetch_run.run_id,
                output_kind="aggregate_eps_parquet",
                path=parquet_path,
                row_count=len(df),
                min_date=min(df["observation_date"]).isoformat() if not df.empty else None,
                max_date=max(df["observation_date"]).isoformat() if not df.empty else None,
                notes="Aggregate EPS workbook snapshots parquet",
            )
            store.record_output(
                run_id=fetch_run.run_id,
                output_kind="aggregate_eps_report",
                path=report_path,
                row_count=len(df),
                min_date=min(df["observation_date"]).isoformat() if not df.empty else None,
                max_date=max(df["observation_date"]).isoformat() if not df.empty else None,
                notes="Aggregate EPS fetch report",
            )
            store.finish_fetch_run(run_id=fetch_run.run_id, status="ok")
        return report_path
    except Exception as exc:
        if store and fetch_run:
            store.finish_fetch_run(run_id=fetch_run.run_id, status="failed", notes=str(exc))
        raise


def _extract_workbook_as_of_date(ws) -> dt.date:
    for row_idx in range(1, min(30, ws.max_row) + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        first = row[0]
        if isinstance(first, dt.datetime):
            return first.date()
    raise AggregateEPSFetchError("Could not find workbook as-of date in ESTIMATES&PEs sheet")


def _parse_legacy_sp500_eps_workbook(workbook_path: Path) -> ParsedAggregateEPSWorkbook:
    df = pd.read_excel(workbook_path, sheet_name=SHEET_NAME, header=None)
    workbook_as_of_date = _extract_legacy_workbook_as_of_date(df)
    table_start_row, header_labels = _find_legacy_observation_header_row(df)

    historical: list[AggregateEPSSnapshot] = []
    current_snapshot: AggregateEPSSnapshot | None = None
    for row_idx in range(table_start_row + 1, len(df)):
        row = df.iloc[row_idx].tolist()
        first = row[0] if row else None
        if isinstance(first, dt.datetime):
            label_map = _build_legacy_observation_value_map(header_labels, row)
            historical.append(
                AggregateEPSSnapshot(
                    observation_date=first.date(),
                    observation_label="historical_quarter_end",
                    forward_estimate_label=_select_legacy_forward_estimate_label(header_labels),
                    forward_estimate_value=_select_legacy_forward_estimate_value(label_map, header_labels),
                    estimate_2025e=None,
                    estimate_q4_2025e=_value_for_legacy_exact_label(label_map, "Q4,'13 EST"),
                    estimate_2026e=None,
                    price=_value_for_legacy_exact_label(label_map, "IDX PRICE"),
                    pe_2025e=None,
                    pe_2026e=None,
                    change_vs_prior_observation_2025e=None,
                    change_vs_prior_observation_q4_2025e=None,
                    change_vs_prior_observation_2026e=None,
                    change_vs_prior_observation_price=None,
                    change_vs_prior_observation_pe_2025e=None,
                    change_vs_prior_observation_pe_2026e=None,
                )
            )
            continue

        if isinstance(first, str) and first.strip().lower() == "current":
            label_map = _build_legacy_observation_value_map(header_labels, row)
            current_snapshot = AggregateEPSSnapshot(
                observation_date=workbook_as_of_date,
                observation_label="current",
                forward_estimate_label=_select_legacy_forward_estimate_label(header_labels),
                forward_estimate_value=_select_legacy_forward_estimate_value(label_map, header_labels),
                estimate_2025e=None,
                estimate_q4_2025e=_value_for_legacy_exact_label(label_map, "Q4,'13 EST"),
                estimate_2026e=None,
                price=_value_for_legacy_exact_label(label_map, "IDX PRICE"),
                pe_2025e=None,
                pe_2026e=None,
                change_vs_prior_observation_2025e=None,
                change_vs_prior_observation_q4_2025e=None,
                change_vs_prior_observation_2026e=None,
                change_vs_prior_observation_price=None,
                change_vs_prior_observation_pe_2025e=None,
                change_vs_prior_observation_pe_2026e=None,
            )
            continue

        if current_snapshot is not None:
            break

    if not historical:
        raise AggregateEPSFetchError("Legacy workbook contained no historical aggregate EPS snapshots")
    if current_snapshot is None:
        raise AggregateEPSFetchError("Legacy workbook missing current aggregate EPS snapshot row")

    return ParsedAggregateEPSWorkbook(
        workbook_as_of_date=workbook_as_of_date,
        public_files_discontinued=False,
        historical_snapshots=historical,
        current_snapshot=current_snapshot,
    )


def _extract_discontinued_flag(ws) -> bool:
    for row_idx in range(1, min(15, ws.max_row) + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        first = row[0]
        if isinstance(first, str) and "public files have been discontinued" in first.lower():
            return True
    return False


def _find_observation_header_row(ws) -> tuple[int, list[str]]:
    for row_idx in range(1, ws.max_row + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        if row[0] == "OBSERVATION":
            labels: list[str] = []
            for value in row[1:]:
                label = str(value).strip() if value is not None else ""
                if label == "OBSERVATION":
                    break
                labels.append(label)
            if any(label.endswith("E") for label in labels):
                return row_idx, labels
    raise AggregateEPSFetchError("Could not find aggregate EPS observation header row")


def _extract_legacy_workbook_as_of_date(df: pd.DataFrame) -> dt.date:
    for row_idx in range(min(10, len(df))):
        value = df.iat[row_idx, 0]
        if isinstance(value, dt.datetime):
            return value.date()
    raise AggregateEPSFetchError("Could not find legacy workbook as-of date in ESTIMATES&PEs sheet")


def _find_legacy_observation_header_row(df: pd.DataFrame) -> tuple[int, list[str]]:
    for row_idx in range(len(df)):
        first = df.iat[row_idx, 0]
        if isinstance(first, str) and first.strip() == "OBSERVATION":
            labels: list[str] = []
            for col_idx in range(1, df.shape[1]):
                value = df.iat[row_idx, col_idx]
                label = str(value).strip() if value is not None else ""
                if not label or label.lower() == "nan":
                    break
                labels.append(label)
            if "2014 EST" in labels or "2013 EST" in labels:
                return row_idx, labels
    raise AggregateEPSFetchError("Could not find legacy aggregate EPS observation header row")


def _build_legacy_observation_value_map(labels: list[str], row: list[object]) -> dict[str, float | None]:
    values: dict[str, float | None] = {}
    for idx, label in enumerate(labels, start=1):
        raw = row[idx] if idx < len(row) else None
        values[label] = _as_float(raw) if raw is not None and not pd.isna(raw) else None
    return values


def _value_for_legacy_exact_label(values: dict[str, float | None], label: str) -> float | None:
    return values.get(label)


def _select_legacy_forward_estimate_label(labels: list[str]) -> str | None:
    for label in reversed(labels):
        if "EST" in label and label != "Q4,'13 EST":
            return label
    return None


def _select_legacy_forward_estimate_value(values: dict[str, float | None], labels: list[str]) -> float | None:
    label = _select_legacy_forward_estimate_label(labels)
    if label is None:
        return None
    return values.get(label)


def _find_current_change_row(ws, header_row: int) -> tuple[float | None, float | None, float | None, float | None, float | None, float | None]:
    for row_idx in range(header_row + 1, min(header_row + 20, ws.max_row) + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        first = row[0]
        if isinstance(first, str) and first.strip().lower() == "change qtr":
            return (
                _as_float(row[1]),
                _as_float(row[2]),
                _as_float(row[3]),
                _as_float(row[4]),
                _as_float(row[5]),
                _as_float(row[6]),
            )
    raise AggregateEPSFetchError("Could not find current aggregate EPS change row")


def _as_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float, str)):
        return float(value)
    raise TypeError(f"_as_float expected int, float, or str, got {type(value).__name__}")


def _build_observation_value_map(labels: list[str], row: tuple[object, ...]) -> dict[str, float | None]:
    values: dict[str, float | None] = {}
    for idx, label in enumerate(labels, start=1):
        if not label:
            continue
        values[label] = _as_float(row[idx]) if idx < len(row) else None
    return values


def _value_for_exact_label(values: dict[str, float | None], label: str) -> float | None:
    return values.get(label)


def _value_for_price(values: dict[str, float | None]) -> float | None:
    return values.get("PRICE") or values.get(" PRICE")


def _value_for_pe(values: dict[str, float | None], year_prefix: str) -> float | None:
    for label, value in values.items():
        normalized = label.replace(" ", "")
        if normalized.startswith(year_prefix) and normalized.endswith("P/E"):
            return value
    return None


def _select_forward_estimate_label(labels: list[str]) -> str | None:
    annual_labels = [
        label
        for label in labels
        if len(label) == 5 and label[:4].isdigit() and label.endswith("E")
    ]
    if not annual_labels:
        return None
    return annual_labels[-1]


def _select_forward_estimate_value(values: dict[str, float | None], labels: list[str]) -> float | None:
    label = _select_forward_estimate_label(labels)
    if label is None:
        return None
    return values.get(label)


