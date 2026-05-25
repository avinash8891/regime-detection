from __future__ import annotations

from collections.abc import Callable
import datetime as dt
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from regime_data_fetch.aggregate_eps_constants import SHEET_NAME
from regime_data_fetch.aggregate_eps_models import (
    AggregateEPSFetchError,
    AggregateEPSSnapshot,
    ParsedAggregateEPSWorkbook,
)


def parse_sp500_eps_workbook(
    workbook_path: Path,
    *,
    read_excel: Callable[..., pd.DataFrame] = pd.read_excel,
) -> ParsedAggregateEPSWorkbook:
    if workbook_path.suffix.lower() == ".xls":
        return _parse_legacy_sp500_eps_workbook(
            workbook_path,
            read_excel=read_excel,
        )

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise AggregateEPSFetchError(f"Workbook missing expected sheet {SHEET_NAME!r}")

    ws = wb[SHEET_NAME]
    workbook_as_of_date = _extract_workbook_as_of_date(ws)
    public_files_discontinued = _extract_discontinued_flag(ws)
    table_start_row, header_labels = _find_observation_header_row(ws)
    current_changes = _find_current_change_row(ws, table_start_row)

    historical: list[AggregateEPSSnapshot] = []
    current_snapshot: AggregateEPSSnapshot | None = None
    for row_idx in range(table_start_row + 1, ws.max_row + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        first = row[0]
        if isinstance(first, dt.datetime):
            label_map = _build_observation_value_map(header_labels, row)
            historical.append(
                AggregateEPSSnapshot(
                    observation_date=first.date(),
                    observation_label="historical_quarter_end",
                    forward_estimate_label=_select_forward_estimate_label(
                        header_labels
                    ),
                    forward_estimate_value=_select_forward_estimate_value(
                        label_map, header_labels
                    ),
                    estimate_2025e=_value_for_exact_label(label_map, "2025E"),
                    estimate_q4_2025e=_value_for_exact_label(label_map, "Q4 2025E"),
                    estimate_2026e=_value_for_exact_label(label_map, "2026E"),
                    price=_value_for_price(label_map),
                    pe_2025e=_value_for_exact_label(label_map, "2025E P/E"),
                    pe_2026e=_value_for_pe(label_map, "2026"),
                    change_vs_prior_observation_2025e=None,
                    change_vs_prior_observation_q4_2025e=None,
                    change_vs_prior_observation_2026e=None,
                    change_vs_prior_observation_price=None,
                    change_vs_prior_observation_pe_2025e=None,
                    change_vs_prior_observation_pe_2026e=None,
                )
            )
            continue

        if isinstance(first, str) and first.strip().lower() == "current":
            label_map = _build_observation_value_map(header_labels, row)
            current_snapshot = AggregateEPSSnapshot(
                observation_date=workbook_as_of_date,
                observation_label="current",
                forward_estimate_label=_select_forward_estimate_label(header_labels),
                forward_estimate_value=_select_forward_estimate_value(
                    label_map, header_labels
                ),
                estimate_2025e=_value_for_exact_label(label_map, "2025E"),
                estimate_q4_2025e=_value_for_exact_label(label_map, "Q4 2025E"),
                estimate_2026e=_value_for_exact_label(label_map, "2026E"),
                price=_value_for_price(label_map),
                pe_2025e=_value_for_exact_label(label_map, "2025E P/E"),
                pe_2026e=_value_for_pe(label_map, "2026"),
                change_vs_prior_observation_2025e=current_changes[0],
                change_vs_prior_observation_q4_2025e=current_changes[1],
                change_vs_prior_observation_2026e=current_changes[2],
                change_vs_prior_observation_price=current_changes[3],
                change_vs_prior_observation_pe_2025e=current_changes[4],
                change_vs_prior_observation_pe_2026e=current_changes[5],
            )
            continue

        if current_snapshot is not None:
            break

    if not historical:
        raise AggregateEPSFetchError(
            "Workbook contained no historical aggregate EPS snapshots"
        )
    if current_snapshot is None:
        raise AggregateEPSFetchError(
            "Workbook missing current aggregate EPS snapshot row"
        )

    return ParsedAggregateEPSWorkbook(
        workbook_as_of_date=workbook_as_of_date,
        public_files_discontinued=public_files_discontinued,
        historical_snapshots=historical,
        current_snapshot=current_snapshot,
    )


def _extract_workbook_as_of_date(ws: Any) -> dt.date:
    for row_idx in range(1, min(30, ws.max_row) + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        first = row[0]
        if isinstance(first, dt.datetime):
            return first.date()
    raise AggregateEPSFetchError(
        "Could not find workbook as-of date in ESTIMATES&PEs sheet"
    )


def _parse_legacy_sp500_eps_workbook(
    workbook_path: Path,
    *,
    read_excel: Callable[..., pd.DataFrame],
) -> ParsedAggregateEPSWorkbook:
    df = read_excel(workbook_path, sheet_name=SHEET_NAME, header=None)
    workbook_as_of_date = _extract_legacy_workbook_as_of_date(df)
    table_start_row, header_labels = _find_legacy_observation_header_row(df)

    historical: list[AggregateEPSSnapshot] = []
    current_snapshot: AggregateEPSSnapshot | None = None
    for row_idx in range(table_start_row + 1, len(df)):
        row = df.iloc[row_idx].tolist()
        first = row[0] if row else None
        if isinstance(first, dt.datetime):
            label_map = _build_legacy_observation_value_map(header_labels, row)
            historical.append(
                AggregateEPSSnapshot(
                    observation_date=first.date(),
                    observation_label="historical_quarter_end",
                    forward_estimate_label=_select_legacy_forward_estimate_label(
                        header_labels
                    ),
                    forward_estimate_value=_select_legacy_forward_estimate_value(
                        label_map, header_labels
                    ),
                    estimate_2025e=None,
                    estimate_q4_2025e=_value_for_legacy_exact_label(
                        label_map, "Q4,'13 EST"
                    ),
                    estimate_2026e=None,
                    price=_value_for_legacy_exact_label(label_map, "IDX PRICE"),
                    pe_2025e=None,
                    pe_2026e=None,
                    change_vs_prior_observation_2025e=None,
                    change_vs_prior_observation_q4_2025e=None,
                    change_vs_prior_observation_2026e=None,
                    change_vs_prior_observation_price=None,
                    change_vs_prior_observation_pe_2025e=None,
                    change_vs_prior_observation_pe_2026e=None,
                )
            )
            continue

        if isinstance(first, str) and first.strip().lower() == "current":
            label_map = _build_legacy_observation_value_map(header_labels, row)
            current_snapshot = AggregateEPSSnapshot(
                observation_date=workbook_as_of_date,
                observation_label="current",
                forward_estimate_label=_select_legacy_forward_estimate_label(
                    header_labels
                ),
                forward_estimate_value=_select_legacy_forward_estimate_value(
                    label_map, header_labels
                ),
                estimate_2025e=None,
                estimate_q4_2025e=_value_for_legacy_exact_label(
                    label_map, "Q4,'13 EST"
                ),
                estimate_2026e=None,
                price=_value_for_legacy_exact_label(label_map, "IDX PRICE"),
                pe_2025e=None,
                pe_2026e=None,
                change_vs_prior_observation_2025e=None,
                change_vs_prior_observation_q4_2025e=None,
                change_vs_prior_observation_2026e=None,
                change_vs_prior_observation_price=None,
                change_vs_prior_observation_pe_2025e=None,
                change_vs_prior_observation_pe_2026e=None,
            )
            continue

        if current_snapshot is not None:
            break

    if not historical:
        raise AggregateEPSFetchError(
            "Legacy workbook contained no historical aggregate EPS snapshots"
        )
    if current_snapshot is None:
        raise AggregateEPSFetchError(
            "Legacy workbook missing current aggregate EPS snapshot row"
        )

    return ParsedAggregateEPSWorkbook(
        workbook_as_of_date=workbook_as_of_date,
        public_files_discontinued=False,
        historical_snapshots=historical,
        current_snapshot=current_snapshot,
    )


def _extract_discontinued_flag(ws: Any) -> bool:
    for row_idx in range(1, min(15, ws.max_row) + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        first = row[0]
        if (
            isinstance(first, str)
            and "public files have been discontinued" in first.lower()
        ):
            return True
    return False


def _find_observation_header_row(ws: Any) -> tuple[int, list[str]]:
    for row_idx in range(1, ws.max_row + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        if row[0] == "OBSERVATION":
            labels: list[str] = []
            for value in row[1:]:
                label = str(value).strip() if value is not None else ""
                if label == "OBSERVATION":
                    break
                labels.append(label)
            if any(label.endswith("E") for label in labels):
                return row_idx, labels
    raise AggregateEPSFetchError("Could not find aggregate EPS observation header row")


def _extract_legacy_workbook_as_of_date(df: pd.DataFrame) -> dt.date:
    for row_idx in range(min(10, len(df))):
        value = df.iat[row_idx, 0]
        if isinstance(value, dt.datetime):
            return value.date()
    raise AggregateEPSFetchError(
        "Could not find legacy workbook as-of date in ESTIMATES&PEs sheet"
    )


def _find_legacy_observation_header_row(df: pd.DataFrame) -> tuple[int, list[str]]:
    for row_idx in range(len(df)):
        first = df.iat[row_idx, 0]
        if isinstance(first, str) and first.strip() == "OBSERVATION":
            labels: list[str] = []
            for col_idx in range(1, df.shape[1]):
                value = df.iat[row_idx, col_idx]
                label = str(value).strip() if value is not None else ""
                if not label or label.lower() == "nan":
                    break
                labels.append(label)
            if "2014 EST" in labels or "2013 EST" in labels:
                return row_idx, labels
    raise AggregateEPSFetchError(
        "Could not find legacy aggregate EPS observation header row"
    )


def _build_legacy_observation_value_map(
    labels: list[str], row: list[object]
) -> dict[str, float | None]:
    values: dict[str, float | None] = {}
    for idx, label in enumerate(labels, start=1):
        raw = row[idx] if idx < len(row) else None
        values[label] = _as_float(raw) if raw is not None and not pd.isna(raw) else None
    return values


def _value_for_legacy_exact_label(
    values: dict[str, float | None], label: str
) -> float | None:
    return values.get(label)


def _select_legacy_forward_estimate_label(labels: list[str]) -> str | None:
    for label in reversed(labels):
        if "EST" in label and label != "Q4,'13 EST":
            return label
    return None


def _select_legacy_forward_estimate_value(
    values: dict[str, float | None], labels: list[str]
) -> float | None:
    label = _select_legacy_forward_estimate_label(labels)
    if label is None:
        return None
    return values.get(label)


def _find_current_change_row(
    ws: Any, header_row: int
) -> tuple[
    float | None, float | None, float | None, float | None, float | None, float | None
]:
    for row_idx in range(header_row + 1, min(header_row + 20, ws.max_row) + 1):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        first = row[0]
        if isinstance(first, str) and first.strip().lower() == "change qtr":
            return (
                _as_float(row[1]),
                _as_float(row[2]),
                _as_float(row[3]),
                _as_float(row[4]),
                _as_float(row[5]),
                _as_float(row[6]),
            )
    raise AggregateEPSFetchError("Could not find current aggregate EPS change row")


def _as_float(value: object) -> float | None:
    if value is None:
        return None
    return float(value)


def _build_observation_value_map(
    labels: list[str], row: tuple[object, ...]
) -> dict[str, float | None]:
    values: dict[str, float | None] = {}
    for idx, label in enumerate(labels, start=1):
        if not label:
            continue
        values[label] = _as_float(row[idx]) if idx < len(row) else None
    return values


def _value_for_exact_label(values: dict[str, float | None], label: str) -> float | None:
    return values.get(label)


def _value_for_price(values: dict[str, float | None]) -> float | None:
    return values.get("PRICE") or values.get(" PRICE")


def _value_for_pe(values: dict[str, float | None], year_prefix: str) -> float | None:
    for label, value in values.items():
        normalized = label.replace(" ", "")
        if normalized.startswith(year_prefix) and normalized.endswith("P/E"):
            return value
    return None


def _select_forward_estimate_label(labels: list[str]) -> str | None:
    annual_labels = [
        label
        for label in labels
        if len(label) == 5 and label[:4].isdigit() and label.endswith("E")
    ]
    if not annual_labels:
        return None
    return annual_labels[-1]


def _select_forward_estimate_value(
    values: dict[str, float | None], labels: list[str]
) -> float | None:
    label = _select_forward_estimate_label(labels)
    if label is None:
        return None
    return values.get(label)
